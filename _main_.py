from openpyxl import load_workbook
import socket
import struct
import ipaddress


def get_all_rows(data_file, ws):

    """Get all rows"""
    xcell = load_workbook(data_file)


    ws = xcell['IPv4']
    all_rows = list(ws.rows)

    return all_rows


'''
def get_all_columns(data_file, ws):
   
    """Get all columns"""
    
    xcell = load_workbook(data_file)
    ws = xcell['IPv4']

    column_a = []
    column_b = []
    column_c = []

    for row in range(2, ws.max_row+1):
        column_a.append(ws.cell(row=row, column=1).value)
        column_b.append(ws.cell(row=row, column=2).value)
        column_c.append(ws.cell(row=row, column=3).value)
        return column_a, column_b, column_c
'''


def summarize_data(data):
    """Summary of addresses, subnet, unicode, and country"""

    ip_addresses = []
    subnets = []
    for row in data[1:220980]:
        if (row[0].value >= 0) and (row[0].value <= 4294967295):
            if row[3].value in ["Russia", "China"]:
                ip_int = row[0].value
                packed_ip = struct.pack("!L", ip_int)
                ip_str = socket.inet_ntoa(packed_ip)
                ip_addresses.append(ip_str)
                sub_int = row[1].value
                packed_sub = struct.pack("!L", sub_int)
                sub_str = socket.inet_ntoa(packed_sub)
                subnets.append(sub_str)
    return ip_addresses, subnets


def user_prompt():
    user_input = input("Enter 'IP' if you'd like to see IPs, or enter 'Subnet' for subnets: ")
    if user_input == 'IP':
        for ip in ip_addresses:
            print(ip)

    elif user_input == "Subnet":
        for sub in subnets:
            print(sub)
    else:
        print("Invalid input. Please enter 'IP' or 'Subnet'")
    


data_file = r'Data\IP2LOCATION-LITE-DB1.xlsx'
data = get_all_rows(data_file, 'IPv4')
column_a, column_b, column_c = get_all_columns(data_file, 'IPv4')

get_all_rows
#get_all_columns
summarize_data(data)
ip_addresses, subnets = summarize_data(data)
user_prompt()
